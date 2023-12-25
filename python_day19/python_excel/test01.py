"""
===============***===============
Auther : Lee
Date   : 2022-11-23 - 16:42
File   : test01.py
===============***===============
"""
import openpyxl

from autotest_03.comms.public_api import get_ini_data

book = openpyxl.load_workbook('D:\WorkSpace\python_day19\python_excel\cases.xlsx')  # 获取excel文件并读取
sheet = book['app']  # 读取文件中表名为app的表
raws = list(sheet.rows)  # 获取excel中的单元格
# print(raws)

# raw = sheet.cell(row=1, column=2).value  # 通过表格中的行数和列数来获取里面的值
# print(len(raw))

# sheet.cell(row=2, column=7, value='tester24班的第一次excel读写!')  # 通过行与列给表中的单元个插入值
# book.save(r'cases.xlsx')


raw1=[]
raw2=[]
for i in range(1,len(raws)+1):
    raw1.append(raw2)
    for j in range(1,len(raws[0])+1):
        raw=sheet.cell(row=i,column=j).value
        raw2.append(raw)
    raw1.append(raw2)
print(raw1)





print(raw2)
print(len(raws))
print(len(raws[0]))